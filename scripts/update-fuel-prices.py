#!/usr/bin/env python3
"""
Weekly fuel price updater.
Sources:
  - EIA.gov        → US national average (requires EIA_API_KEY env var)
  - EU Oil Bulletin → All 27 EU countries via stable EC download URL (no key)
  - Everything else → left unchanged from previous week

Usage:
  EIA_API_KEY=your_key python3 scripts/update-fuel-prices.py
"""

import json
import os
import sys
import io
import requests
import openpyxl
from datetime import date

PRICES_FILE = "fuel-prices.json"
TIMEOUT = 20

# Stable URL — always serves the latest weekly bulletin (prices WITH taxes)
EU_BULLETIN_URL = (
    "https://energy.ec.europa.eu/document/download/"
    "264c2d0f-f161-4ea3-a777-78faae59bea0_en"
)

# Rows 3–29 in the XLSX map to EU countries in this order (alphabetical)
EU_ROW_TO_CODE = {
    3:  "BE", 4:  "BG", 5:  "CZ", 6:  "DK", 7:  "DE",
    8:  "EE", 9:  "IE", 10: "GR", 11: "ES", 12: "FR",
    13: "HR", 14: "IT", 15: "CY", 16: "LV", 17: "LT",
    18: "LU", 19: "HU", 20: "MT", 21: "NL", 22: "AT",
    23: "PL", 24: "PT", 25: "RO", 26: "SI", 27: "SK",
    28: "FI", 29: "SE",
}


def load_prices():
    with open(PRICES_FILE, "r") as f:
        return json.load(f)


def save_prices(data):
    with open(PRICES_FILE, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    print(f"✓ Saved {PRICES_FILE}")


def fetch_us_prices(api_key):
    """Returns (petrol_per_litre, diesel_per_litre) in USD, or (None, None)."""
    base = (
        "https://api.eia.gov/v2/petroleum/pri/gnd/data/"
        "?api_key={key}&frequency=weekly&data[0]=value"
        "&facets[duoarea][]=NUS&facets[product][]={product}"
        "&sort[0][column]=period&sort[0][direction]=desc&offset=0&length=1"
    )
    results = {}
    for label, product in [("petrol", "EPM0"), ("diesel", "EPD2D")]:
        try:
            r = requests.get(base.format(key=api_key, product=product), timeout=TIMEOUT)
            r.raise_for_status()
            per_gallon = float(r.json()["response"]["data"][0]["value"])
            per_litre  = round(per_gallon / 3.78541, 4)
            print(f"  US {label}: ${per_gallon:.3f}/gal → ${per_litre:.4f}/L")
            results[label] = per_litre
        except Exception as e:
            print(f"  ✗ EIA {label} failed: {e}")
    return results.get("petrol"), results.get("diesel")


def fetch_eu_prices():
    """
    Downloads the EU Oil Bulletin XLSX and extracts petrol + diesel prices.
    Prices in the file are EUR per 1000 litres → divide by 1000 for per-litre.
    Returns dict { "DE": {"petrol": 1.75, "diesel": 1.68}, ... }
    """
    try:
        r = requests.get(EU_BULLETIN_URL, timeout=TIMEOUT)
        r.raise_for_status()

        wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
        ws = wb.active

        results = {}
        for row_idx, code in EU_ROW_TO_CODE.items():
            try:
                petrol_raw = ws.cell(row=row_idx, column=2).value  # Col B: Euro-super 95
                diesel_raw = ws.cell(row=row_idx, column=3).value  # Col C: Automotive gas oil
                if petrol_raw is None or diesel_raw is None:
                    continue
                petrol = round(float(petrol_raw) / 1000, 4)  # EUR/1000L → EUR/L
                diesel = round(float(diesel_raw) / 1000, 4)
                results[code] = {"petrol": petrol, "diesel": diesel}
                print(f"  {code}: petrol={petrol:.4f} diesel={diesel:.4f} EUR/L")
            except Exception as e:
                print(f"  ✗ {code} row {row_idx}: {e}")

        return results

    except Exception as e:
        print(f"  ✗ EU Oil Bulletin failed: {e}")
        return {}


def main():
    print(f"\n{'='*50}")
    print(f"Fuel price update — {date.today()}")
    print(f"{'='*50}\n")

    prices = load_prices()
    updated = []

    # ── US via EIA ────────────────────────────────────────────────────────────
    eia_key = os.environ.get("EIA_API_KEY", "").strip()
    if eia_key:
        print("→ Fetching US prices from EIA...")
        petrol, diesel = fetch_us_prices(eia_key)
        if petrol is not None:
            prices["countries"]["US"]["petrol"] = petrol
            updated.append("US petrol")
        if diesel is not None:
            prices["countries"]["US"]["diesel"] = diesel
            updated.append("US diesel")
    else:
        print("→ EIA_API_KEY not set — skipping US")

    # ── EU via Oil Bulletin ───────────────────────────────────────────────────
    print("\n→ Fetching EU prices from EU Oil Bulletin...")
    for code, vals in fetch_eu_prices().items():
        if code in prices["countries"]:
            prices["countries"][code]["petrol"] = vals["petrol"]
            prices["countries"][code]["diesel"] = vals["diesel"]
            updated.append(code)

    # ── Wrap up ───────────────────────────────────────────────────────────────
    prices["_meta"]["updated"] = str(date.today())

    print(f"\n{'='*50}")
    if updated:
        print(f"✓ Updated: {', '.join(updated)}")
    else:
        print("⚠ No prices updated — all sources failed or no API key.")
    print(f"{'='*50}\n")

    save_prices(prices)
    return 0


if __name__ == "__main__":
    sys.exit(main())
